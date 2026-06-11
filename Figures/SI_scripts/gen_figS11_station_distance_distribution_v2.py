from __future__ import annotations

import math
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook

from paper_plot_style import PALETTE, add_panel_label, apply_style, save_figure


ROOT = Path(__file__).resolve().parent
WORKSPACE_ROOT = ROOT.parent.parent
FOR_REVIEW = WORKSPACE_ROOT / "for_review"
EARTH_RADIUS_KM = 6371.0088


def find_station_file() -> Path:
    candidates = sorted(WORKSPACE_ROOT.glob("*WGS1984*.xlsx"))
    if not candidates:
        raise FileNotFoundError("No station workbook matching '*WGS1984*.xlsx' was found in for_review.")
    return candidates[0]


def is_number(value: object) -> bool:
    try:
        float(value)
        return True
    except (TypeError, ValueError):
        return False


def haversine_km(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)
    a = (
        math.sin(dphi / 2.0) ** 2
        + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2.0) ** 2
    )
    return 2.0 * EARTH_RADIUS_KM * math.asin(math.sqrt(a))


def load_stations() -> list[dict[str, object]]:
    workbook = load_workbook(find_station_file(), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    first_data_row = next(
        (row for row in rows[1:] if any(cell is not None and str(cell).strip() for cell in row)),
        None,
    )
    if first_data_row is None:
        return []

    legacy_layout = is_number(first_data_row[0]) and is_number(first_data_row[1])

    stations: list[dict[str, object]] = []
    for row in rows[1:]:
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue

        if legacy_layout:
            name_idx, city_idx, platform_idx = 2, 3, 4
            lon = row[7] if len(row) > 7 else None
            lat = row[8] if len(row) > 8 else None
            if lon is None or lat is None:
                lon = row[1] if len(row) > 1 else None
                lat = row[0] if len(row) > 0 else None
        else:
            name_idx, city_idx, platform_idx = 0, 1, 2
            lon = row[5] if len(row) > 5 else None
            lat = row[6] if len(row) > 6 else None

        if lon is None or lat is None:
            continue

        stations.append(
            {
                "name": str(row[name_idx]).strip(),
                "city": str(row[city_idx]).strip(),
                "platform": str(row[platform_idx]).strip(),
                "lon": float(lon),
                "lat": float(lat),
            }
        )
    return stations


def pairwise_distances(stations: list[dict[str, object]]) -> np.ndarray:
    distances: list[float] = []
    for i in range(len(stations)):
        s1 = stations[i]
        for j in range(i + 1, len(stations)):
            s2 = stations[j]
            distances.append(
                haversine_km(
                    float(s1["lat"]),
                    float(s1["lon"]),
                    float(s2["lat"]),
                    float(s2["lon"]),
                )
            )
    return np.asarray(distances, dtype=float)


def close_pairs(
    stations: list[dict[str, object]], threshold_km: float
) -> list[tuple[float, dict[str, object], dict[str, object]]]:
    pairs: list[tuple[float, dict[str, object], dict[str, object]]] = []
    for i in range(len(stations)):
        s1 = stations[i]
        for j in range(i + 1, len(stations)):
            s2 = stations[j]
            distance = haversine_km(
                float(s1["lat"]),
                float(s1["lon"]),
                float(s2["lat"]),
                float(s2["lon"]),
            )
            if distance < threshold_km:
                pairs.append((distance, s1, s2))
    pairs.sort(key=lambda item: item[0])
    return pairs


def add_summary_box(ax: plt.Axes, distances: np.ndarray) -> None:
    summary = (
        f"Pairs = {len(distances):,}\n"
        f"Mean = {distances.mean():.1f} km\n"
        f"Median = {np.median(distances):.1f} km\n"
        f"Min = {distances.min():.3f} km\n"
        f"Max = {distances.max():.1f} km"
    )
    ax.text(
        0.98,
        0.96,
        summary,
        transform=ax.transAxes,
        ha="right",
        va="top",
        fontsize=8.1,
        bbox={
            "boxstyle": "round,pad=0.28",
            "facecolor": "white",
            "edgecolor": PALETTE["grey_light"],
            "linewidth": 0.8,
        },
    )


def plot_main_histogram(ax: plt.Axes, distances: np.ndarray, bins: np.ndarray) -> None:
    ax.hist(
        distances,
        bins=bins,
        color=PALETTE["teal"],
        edgecolor="white",
        linewidth=0.7,
        alpha=0.9,
    )
    ax.axvspan(0.0, 0.3, color=PALETTE["orange"], alpha=0.12, zorder=0)
    ax.axvline(
        distances.mean(),
        color=PALETTE["red"],
        linestyle="--",
        linewidth=1.2,
        label=f"Mean ({distances.mean():.1f} km)",
    )
    ax.axvline(
        float(np.median(distances)),
        color=PALETTE["gold"],
        linestyle=":",
        linewidth=1.3,
        label=f"Median ({np.median(distances):.1f} km)",
    )
    ax.set_ylabel("Count")
    ax.set_title("All Shenzhen Station Pairs")
    ax.grid(axis="y")
    ax.legend(frameon=False, loc="upper left")
    add_summary_box(ax, distances)


def plot_zoom_histogram(
    ax: plt.Axes,
    distances: np.ndarray,
    bins: np.ndarray,
    close_pair_count: int,
    nearest_pair_m: float,
) -> None:
    ax.hist(
        distances,
        bins=bins,
        color=PALETTE["blue"],
        edgecolor="white",
        linewidth=0.7,
        alpha=0.92,
    )
    ax.axvspan(0.0, 0.3, color=PALETTE["orange"], alpha=0.18, zorder=0)
    ax.axvline(0.3, color=PALETTE["orange"], linestyle="--", linewidth=1.1)
    ax.set_xlim(0.0, float(bins[-1]))
    ax.set_xlabel("Station-pair distance (km)")
    ax.set_ylabel("Count")
    ax.set_title("Near-Zero Distance Zoom (0-2 km)")
    ax.grid(axis="y")
    ax.text(
        0.98,
        0.93,
        f"<300 m: {close_pair_count} pair\nNearest pair = {nearest_pair_m:.1f} m",
        transform=ax.transAxes,
        ha="right",
        va="top",
        fontsize=8.0,
        bbox={
            "boxstyle": "round,pad=0.28",
            "facecolor": "white",
            "edgecolor": PALETTE["orange"],
            "linewidth": 0.8,
        },
    )


def main() -> None:
    apply_style()

    stations = load_stations()
    shenzhen = [station for station in stations if station["city"] == "深圳"]
    if len(shenzhen) < 2:
        raise ValueError("At least two Shenzhen stations are required to plot pairwise distances.")

    shenzhen_distances = pairwise_distances(shenzhen)
    sub_300m_pairs = close_pairs(shenzhen, threshold_km=0.3)
    nearest_pair_m = sub_300m_pairs[0][0] * 1000.0 if sub_300m_pairs else float("nan")

    fig, (ax_main, ax_zoom) = plt.subplots(
        2,
        1,
        figsize=(6.4, 5.6),
        gridspec_kw={"height_ratios": [2.4, 1.3], "hspace": 0.34},
    )

    main_bins = np.linspace(0.0, max(45.0, float(shenzhen_distances.max()) * 1.02), 26)
    zoom_bins = np.linspace(0.0, 2.0, 21)

    plot_main_histogram(ax_main, shenzhen_distances, main_bins)
    ax_main.set_xlim(0.0, float(main_bins[-1]))
    add_panel_label(ax_main, "A")

    plot_zoom_histogram(
        ax_zoom,
        shenzhen_distances,
        zoom_bins,
        close_pair_count=len(sub_300m_pairs),
        nearest_pair_m=nearest_pair_m,
    )
    add_panel_label(ax_zoom, "B")

    fig.suptitle(
        f"Distribution of Pairwise Distances Between UAV Stations in Shenzhen ({len(shenzhen)} stations)",
        y=0.98,
    )
    fig.subplots_adjust(left=0.11, right=0.98, bottom=0.09, top=0.88)
    save_figure(fig, "S11_station_distance_distribution")
    plt.close(fig)


if __name__ == "__main__":
    main()
