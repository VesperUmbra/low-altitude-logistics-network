from __future__ import annotations

import json
import math
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
WORKSPACE_ROOT = ROOT.parent.parent
RAW_CSV = WORKSPACE_ROOT / "for_review" / "ods_sq_flight_dynamic_data.csv"
STATION_XLSX = next(WORKSPACE_ROOT.glob("*WGS1984.xlsx"))

REVIEW_DATA_DIR = WORKSPACE_ROOT / "for_review" / "review_data"
SOURCE_JSON_DIR = REVIEW_DATA_DIR / "source_json"
RANKED_CELL_TABLE = REVIEW_DATA_DIR / "ranked_cell_traffic_table.csv"
HOTSPOT_TABLE = REVIEW_DATA_DIR / "hotspot_exceedance_rank_table.csv"
BREAKPOINT_JSON = SOURCE_JSON_DIR / "breakpoint_results.json"

OUTPUT_JSON = SOURCE_JSON_DIR / "real_station_anchor_results.json"
OUTPUT_CSV = REVIEW_DATA_DIR / "real_station_anchor_sites_mapped.csv"
GRID_INFO_JSON = ROOT / "data" / "processed" / "full_100m" / "grid" / "grid_info.json"

GRID_SIZE_M = 100.0
MIN_LON = 113.31
MAX_LON = 114.25
MIN_LAT = 22.31
MAX_LAT = 22.87
LAT_CENTER = (MIN_LAT + MAX_LAT) / 2
METERS_PER_DEG_LAT = 111000.0
METERS_PER_DEG_LON = 111000.0 * math.cos(math.radians(LAT_CENTER))
N_ROWS = math.ceil(((MAX_LAT - MIN_LAT) * METERS_PER_DEG_LAT) / GRID_SIZE_M)
N_COLS = math.ceil(((MAX_LON - MIN_LON) * METERS_PER_DEG_LON) / GRID_SIZE_M)
TIME_WINDOW_S = int(json.loads(GRID_INFO_JSON.read_text(encoding="utf-8")).get("time_window_s", 300))
WINDOW_FLOOR_ALIAS = f"{TIME_WINDOW_S // 60}min" if TIME_WINDOW_S % 60 == 0 else f"{TIME_WINDOW_S}s"

USECOLS = ["order_id", "time", "altitude", "speed", "longitude", "latitude", "flight_time"]
CHUNK_SIZE = 250_000


def relative_risk_ci(a: int, b: int, c: int, d: int) -> tuple[float, float, float]:
    rr = (a / (a + b)) / (c / (c + d))
    se = math.sqrt((1 / a) - (1 / (a + b)) + (1 / c) - (1 / (c + d)))
    delta = 1.96 * se
    return rr, math.exp(math.log(rr) - delta), math.exp(math.log(rr) + delta)


def map_to_grid(lon: pd.Series, lat: pd.Series) -> tuple[pd.Series, pd.Series]:
    rows = (((lat - MIN_LAT) * METERS_PER_DEG_LAT) / GRID_SIZE_M).astype(int)
    cols = (((lon - MIN_LON) * METERS_PER_DEG_LON) / GRID_SIZE_M).astype(int)
    rows = rows.clip(lower=0, upper=N_ROWS - 1)
    cols = cols.clip(lower=0, upper=N_COLS - 1)
    return rows.astype("int16"), cols.astype("int16")


def expand_buffer(cell_ids: set[str], radius: int) -> set[str]:
    expanded: set[str] = set()
    for cell_id in cell_ids:
        row_str, col_str = cell_id.split("_")
        row = int(row_str)
        col = int(col_str)
        for dr in range(-radius, radius + 1):
            for dc in range(-radius, radius + 1):
                rr = row + dr
                cc = col + dc
                if 0 <= rr < N_ROWS and 0 <= cc < N_COLS:
                    expanded.add(f"{rr}_{cc}")
    return expanded


def normalize_open_time(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def load_station_rows() -> list[dict[str, object]]:
    wb = load_workbook(STATION_XLSX, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    items: list[dict[str, object]] = []
    for raw in rows[1:]:
        if not any(v is not None for v in raw):
            continue
        row = {headers[i]: raw[i] for i in range(len(headers))}
        items.append(row)
    return items


def parse_station_inventory() -> dict[str, pd.DataFrame]:
    rows = load_station_rows()
    df = pd.DataFrame(rows)
    df = df[df["城市"].astype(str) == "深圳"].copy()

    df["platform"] = df["平台"].astype(str)
    df["site_name"] = df["站点名"].astype(str)
    df["note"] = df["备注"].fillna("").astype(str)
    df["open_time"] = df["开通时间"].apply(normalize_open_time)
    df["lon"] = pd.to_numeric(df["WGS1984_经度"], errors="coerce")
    df["lat"] = pd.to_numeric(df["WGS1984_纬度"], errors="coerce")
    df = df[df["lon"].notna() & df["lat"].notna()].copy()

    # Keep every Shenzhen site in the main set as requested, and expose a
    # conservative sensitivity set that removes explicit test stations.
    df["is_test_site"] = df["note"].str.contains("测试", na=False) | df["site_name"].str.contains("测试", na=False)
    rows_mapped, cols_mapped = map_to_grid(df["lon"], df["lat"])
    df["row"] = rows_mapped.astype(int)
    df["col"] = cols_mapped.astype(int)
    df["cell_id"] = df["row"].astype(str) + "_" + df["col"].astype(str)

    operational = df[~df["is_test_site"]].copy()
    return {"all_shenzhen_sites": df, "operational_sites": operational}


@dataclass
class AggregatedCounts:
    total_points_by_cell: Counter
    total_samples_by_cell: Counter
    exceedance_samples_by_cell: Counter
    endpoint_cells: set[str]
    total_exceedance_samples: int
    cleaned_points: int
    valid_orders: int


def clean_chunk(chunk: pd.DataFrame) -> pd.DataFrame:
    chunk["time"] = pd.to_numeric(chunk["time"], errors="coerce")
    chunk["speed"] = pd.to_numeric(chunk["speed"], errors="coerce")
    chunk["altitude"] = pd.to_numeric(chunk["altitude"], errors="coerce")
    chunk["longitude"] = pd.to_numeric(chunk["longitude"], errors="coerce")
    chunk["latitude"] = pd.to_numeric(chunk["latitude"], errors="coerce")
    chunk["flight_time"] = pd.to_numeric(chunk["flight_time"], errors="coerce")

    valid = (
        chunk["time"].notna()
        & chunk["speed"].between(0, 40, inclusive="both")
        & chunk["altitude"].between(0, 500, inclusive="both")
        & chunk["longitude"].between(MIN_LON, MAX_LON, inclusive="both")
        & chunk["latitude"].between(MIN_LAT, MAX_LAT, inclusive="both")
        & (chunk["flight_time"] >= 0)
    )
    chunk = chunk.loc[valid].copy()
    if chunk.empty:
        return chunk

    chunk["datetime"] = pd.to_datetime(chunk["time"].astype("int64").astype(str), format="%Y%m%d%H%M%S", errors="coerce")
    chunk = chunk[chunk["datetime"].notna()].copy()
    if chunk.empty:
        return chunk

    rows, cols = map_to_grid(chunk["longitude"], chunk["latitude"])
    chunk["row"] = rows.astype(int)
    chunk["col"] = cols.astype(int)
    chunk["cell_id"] = chunk["row"].astype(str) + "_" + chunk["col"].astype(str)
    chunk["window_start"] = chunk["datetime"].dt.floor(WINDOW_FLOOR_ALIAS)
    return chunk


def collect_valid_orders() -> set[str]:
    order_counts: Counter[str] = Counter()
    for chunk in pd.read_csv(RAW_CSV, usecols=USECOLS, chunksize=CHUNK_SIZE, low_memory=False):
        chunk = clean_chunk(chunk)
        if chunk.empty:
            continue
        order_counts.update(chunk.groupby("order_id", observed=True).size().to_dict())
    return {str(order_id) for order_id, count in order_counts.items() if int(count) >= 3}


def aggregate_from_raw_csv(rho_star: float) -> AggregatedCounts:
    total_points_by_cell: Counter[str] = Counter()
    total_samples_by_cell: Counter[str] = Counter()
    exceedance_samples_by_cell: Counter[str] = Counter()
    spatiotemporal_point_counts: Counter[tuple[str, str]] = Counter()
    first_by_order: dict[str, tuple[int, str]] = {}
    last_by_order: dict[str, tuple[int, str]] = {}
    cleaned_points = 0
    valid_orders = collect_valid_orders()

    for chunk in pd.read_csv(RAW_CSV, usecols=USECOLS, chunksize=CHUNK_SIZE, low_memory=False):
        chunk = clean_chunk(chunk)
        if chunk.empty:
            continue

        chunk["order_id"] = chunk["order_id"].astype(str)
        chunk = chunk[chunk["order_id"].isin(valid_orders)].copy()
        if chunk.empty:
            continue

        cleaned_points += int(len(chunk))

        cell_counts = chunk.groupby("cell_id", observed=True).size()
        total_points_by_cell.update(cell_counts.to_dict())

        st_counts = (
            chunk.groupby(["cell_id", "window_start"], observed=True)
            .size()
            .reset_index(name="n_points")
        )
        for row in st_counts.itertuples(index=False):
            key = (str(row.cell_id), pd.Timestamp(row.window_start).isoformat())
            spatiotemporal_point_counts[key] += int(row.n_points)

        first_chunk = (
            chunk.sort_values(["order_id", "time"])
            .groupby("order_id", sort=False, observed=True)
            .first()[["time", "cell_id"]]
        )
        for order_id, row in first_chunk.iterrows():
            time_value = int(row["time"])
            cell_id = str(row["cell_id"])
            prev = first_by_order.get(str(order_id))
            if prev is None or time_value < prev[0]:
                first_by_order[str(order_id)] = (time_value, cell_id)

        last_chunk = (
            chunk.sort_values(["order_id", "time"])
            .groupby("order_id", sort=False, observed=True)
            .last()[["time", "cell_id"]]
        )
        for order_id, row in last_chunk.iterrows():
            time_value = int(row["time"])
            cell_id = str(row["cell_id"])
            prev = last_by_order.get(str(order_id))
            if prev is None or time_value > prev[0]:
                last_by_order[str(order_id)] = (time_value, cell_id)

    total_exceedance_samples = 0
    for (cell_id, _window_start), n_points in spatiotemporal_point_counts.items():
        total_samples_by_cell[cell_id] += 1
        if n_points >= rho_star:
            exceedance_samples_by_cell[cell_id] += 1
            total_exceedance_samples += 1

    endpoint_cells = {cell_id for _, cell_id in first_by_order.values()} | {cell_id for _, cell_id in last_by_order.values()}
    return AggregatedCounts(
        total_points_by_cell=total_points_by_cell,
        total_samples_by_cell=total_samples_by_cell,
        exceedance_samples_by_cell=exceedance_samples_by_cell,
        endpoint_cells=endpoint_cells,
        total_exceedance_samples=total_exceedance_samples,
        cleaned_points=cleaned_points,
        valid_orders=len(valid_orders),
    )


def build_sets(aggregated: AggregatedCounts) -> dict[str, set[str]]:
    active_cells = set(aggregated.total_points_by_cell)
    sorted_cells = sorted(aggregated.total_points_by_cell.items(), key=lambda item: item[1], reverse=True)
    n_backbone = max(1, int(len(sorted_cells) * 0.10))
    backbone_cells = {cell_id for cell_id, _ in sorted_cells[:n_backbone]}
    hotspot_cells = set(aggregated.exceedance_samples_by_cell)
    return {
        "active_cells": active_cells,
        "backbone_cells": backbone_cells,
        "hotspot_cells": hotspot_cells,
        "endpoint_cells": aggregated.endpoint_cells,
    }


def summarize_anchor(
    anchor_cells: set[str],
    aggregated: AggregatedCounts,
    active_cells: set[str],
    backbone_cells: set[str],
    hotspot_cells: set[str],
    endpoint_cells: set[str],
) -> dict[str, object]:
    active_in = len(active_cells & anchor_cells)
    backbone_in = len(backbone_cells & anchor_cells)
    hotspot_in = len(hotspot_cells & anchor_cells)
    endpoint_in = len(endpoint_cells & anchor_cells)

    a = sum(aggregated.exceedance_samples_by_cell.get(cell_id, 0) for cell_id in anchor_cells)
    anchor_total_samples = sum(aggregated.total_samples_by_cell.get(cell_id, 0) for cell_id in anchor_cells)
    b = anchor_total_samples - a
    total_samples = sum(aggregated.total_samples_by_cell.values())
    c = aggregated.total_exceedance_samples - a
    d = (total_samples - anchor_total_samples) - c
    rr, rr_lo, rr_hi = relative_risk_ci(a, b, c, d)

    return {
        "anchor_cells": int(len(anchor_cells)),
        "active_cells_covered": int(active_in),
        "active_cells_covered_share": float(active_in / len(active_cells)) if active_cells else 0.0,
        "backbone_cells_covered": int(backbone_in),
        "backbone_cells_covered_share": float(backbone_in / len(backbone_cells)) if backbone_cells else 0.0,
        "hotspot_cells_covered": int(hotspot_in),
        "hotspot_cells_covered_share": float(hotspot_in / len(hotspot_cells)) if hotspot_cells else 0.0,
        "endpoint_cells_covered": int(endpoint_in),
        "endpoint_cells_covered_share": float(endpoint_in / len(endpoint_cells)) if endpoint_cells else 0.0,
        "exceedance_samples_covered": int(a),
        "exceedance_samples_covered_share": float(a / aggregated.total_exceedance_samples)
        if aggregated.total_exceedance_samples
        else 0.0,
        "sample_rr_exceedance": float(rr),
        "sample_rr_exceedance_ci": [float(rr_lo), float(rr_hi)],
        "contingency": {
            "anchor_and_exceedance": int(a),
            "anchor_and_non_exceedance": int(b),
            "non_anchor_and_exceedance": int(c),
            "non_anchor_and_non_exceedance": int(d),
        },
    }


def station_metadata(df: pd.DataFrame) -> dict[str, object]:
    return {
        "site_rows": int(len(df)),
        "unique_site_names": int(df["site_name"].nunique()),
        "unique_anchor_cells": int(df["cell_id"].nunique()),
        "platform_counts": {str(k): int(v) for k, v in df["platform"].value_counts().sort_index().items()},
        "test_site_rows": int(df["is_test_site"].sum()),
        "open_time_examples": sorted({s for s in df["open_time"] if s})[:10],
    }


def main() -> None:
    with open(BREAKPOINT_JSON, "r", encoding="utf-8") as f:
        rho_star = float(json.load(f)["consensus_rho_star"])

    station_sets = parse_station_inventory()
    aggregated = aggregate_from_raw_csv(rho_star=rho_star)
    cell_sets = build_sets(aggregated)

    results: dict[str, object] = {
        "anchor_name": "real_uav_station_anchor",
        "description": "Externally supplied Shenzhen UAV station inventory mapped to the manuscript's 100 m grid.",
        "query_timestamp_utc": datetime.now(UTC).isoformat(),
        "rho_star_used": rho_star,
        "raw_csv": str(RAW_CSV),
        "station_workbook": STATION_XLSX.name,
        "counts_from_raw": {
            "cleaned_points": int(aggregated.cleaned_points),
            "valid_orders": int(aggregated.valid_orders),
            "active_cells": int(len(cell_sets["active_cells"])),
            "backbone_cells": int(len(cell_sets["backbone_cells"])),
            "hotspot_cells": int(len(cell_sets["hotspot_cells"])),
            "endpoint_cells": int(len(cell_sets["endpoint_cells"])),
            "total_exceedance_samples": int(aggregated.total_exceedance_samples),
            "total_cell_window_samples": int(sum(aggregated.total_samples_by_cell.values())),
        },
        "anchor_sets": {},
    }

    mapped_rows: list[dict[str, object]] = []
    for anchor_name, df in station_sets.items():
        exact_cells = set(df["cell_id"].astype(str))
        buffer1 = expand_buffer(exact_cells, radius=1)
        buffer2 = expand_buffer(exact_cells, radius=2)

        results["anchor_sets"][anchor_name] = {
            "station_metadata": station_metadata(df),
            "exact_anchor_cells": summarize_anchor(
                exact_cells,
                aggregated,
                cell_sets["active_cells"],
                cell_sets["backbone_cells"],
                cell_sets["hotspot_cells"],
                cell_sets["endpoint_cells"],
            ),
            "buffer1_anchor_cells": summarize_anchor(
                buffer1,
                aggregated,
                cell_sets["active_cells"],
                cell_sets["backbone_cells"],
                cell_sets["hotspot_cells"],
                cell_sets["endpoint_cells"],
            ),
            "buffer2_anchor_cells": summarize_anchor(
                buffer2,
                aggregated,
                cell_sets["active_cells"],
                cell_sets["backbone_cells"],
                cell_sets["hotspot_cells"],
                cell_sets["endpoint_cells"],
            ),
        }

        for row in df.itertuples(index=False):
            mapped_rows.append(
                {
                    "anchor_set": anchor_name,
                    "site_name": row.site_name,
                    "platform": row.platform,
                    "note": row.note,
                    "open_time": row.open_time,
                    "lon": row.lon,
                    "lat": row.lat,
                    "grid_row": int(row.row),
                    "grid_col": int(row.col),
                    "grid_cell_id": row.cell_id,
                    "is_test_site": bool(row.is_test_site),
                    "is_active_cell": row.cell_id in cell_sets["active_cells"],
                    "is_backbone_cell": row.cell_id in cell_sets["backbone_cells"],
                    "is_hotspot_cell": row.cell_id in cell_sets["hotspot_cells"],
                    "is_endpoint_cell": row.cell_id in cell_sets["endpoint_cells"],
                }
            )

    OUTPUT_JSON.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")
    pd.DataFrame(mapped_rows).to_csv(OUTPUT_CSV, index=False, encoding="utf-8-sig")

    print(f"Wrote {OUTPUT_JSON}")
    print(f"Wrote {OUTPUT_CSV}")
    for anchor_name, payload in results["anchor_sets"].items():
        exact = payload["exact_anchor_cells"]
        buffer1 = payload["buffer1_anchor_cells"]
        print(
            anchor_name,
            f"sites={payload['station_metadata']['site_rows']}",
            f"exact_active={exact['active_cells_covered_share']:.4f}",
            f"exact_hotspot={exact['hotspot_cells_covered_share']:.4f}",
            f"buffer1_hotspot={buffer1['hotspot_cells_covered_share']:.4f}",
            f"buffer1_exceedance={buffer1['exceedance_samples_covered_share']:.4f}",
            f"buffer1_rr={buffer1['sample_rr_exceedance']:.2f}",
        )


if __name__ == "__main__":
    main()
